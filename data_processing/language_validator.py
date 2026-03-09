"""
Created on Mon May 27 09:06:17 2024

@author: antonioyepez
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from langdetect import detect, LangDetectException

# Language mapping
LANGUAGE_MAP = {
    'cs': 'cs',
    'de': 'de',
    'en': 'en',
    'es': 'es',
    'fr': 'fr',
    'it': 'it',
    'nl': 'nl',
    'pl': 'pl',
    'pt': 'pt'
}

# Visual feedback for editors
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

input_dir = "/Users/antonioyepez/Desktop/Scripts/Recetas/Recetas_por_idioma"
output_dir = "/Users/antonioyepez/Desktop/Scripts/Recetas/Recetas_pendientes"

if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Attemps to identify t he language of the provided text
def detect_language(text):
    try:
        return detect(text)
    except LangDetectException:
        return None
# Scans excel files to detect language mismatches or empty fields
def process_file(file_path, language_code):
    wb = load_workbook(file_path)
    ws = wb.active

    rows_to_save = []
    header = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, max_col=20):
        row_data = [cell.value for cell in row]
        original_lang = row_data[1]
        errors = False

        for col_index, cell in enumerate(row[1:], start=1):
            cell_value = cell.value
            if cell_value:
                detected_lang = detect_language(cell_value)
                if detected_lang and detected_lang != language_code:
                    errors = True
                    cell.fill = RED_FILL
            else:
                errors = True
                cell.fill = RED_FILL

        if errors:
            rows_to_save.append(row_data)
    # Exports a dedicated review report
    if rows_to_save:
        output_file = os.path.join(output_dir, f"Pending_Recipes_{language_code}.xlsx")
        save_to_excel(header, rows_to_save, output_file)
    # Save the original file with the markers
    wb.save(file_path)

def save_to_excel(header, data, file_path):
    df = pd.DataFrame(data, columns=header)
    df.to_excel(file_path, index=False)

for file_name in os.listdir(input_dir):
    if file_name.endswith(".xlsx"):
        file_path = os.path.join(input_dir, file_name)
        language_code = file_name.split('_')[-1].split('.')[0]
        if language_code in LANGUAGE_MAP:
            process_file(file_path, LANGUAGE_MAP[language_code])

print("Process completed.")
